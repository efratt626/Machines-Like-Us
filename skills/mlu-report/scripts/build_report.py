"""
MLU Report Builder
Usage: python3 build_report.py <root_folder> <period> <report_type>
  root_folder  – path containing Apple/, Spotify/, YouTube/ subfolders
  period       – e.g. "Q1 2026" or "March 2026"
  report_type  – "QBR" or "Monthly"
"""

import sys
import os
import glob
import shutil
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

ROOT        = sys.argv[1]
PERIOD      = sys.argv[2]   # e.g. "Q1 2026" or "March 2026"
REPORT_TYPE = sys.argv[3]   # "QBR" or "Monthly"
IMG_DIR     = os.path.join(ROOT, "charts_tmp")
os.makedirs(IMG_DIR, exist_ok=True)

# ── Validation ─────────────────────────────────────────────────────────────

def check(path, label):
    if not os.path.exists(path):
        print(f"ERROR: Missing {label}: {path}")
        sys.exit(1)

check(os.path.join(ROOT, "Apple", "Apple Monthly Listeners"),   "Apple/Apple Monthly Listeners/")
check(os.path.join(ROOT, "Apple", "Apple Monthly Subscribers"), "Apple/Apple Monthly Subscribers/")
check(os.path.join(ROOT, "Spotify"),                            "Spotify/")
check(os.path.join(ROOT, "YouTube"),                            "YouTube/")

def glob_one(pattern, label):
    matches = glob.glob(pattern)
    if not matches:
        print(f"ERROR: No file found for {label} (pattern: {pattern})")
        sys.exit(1)
    return matches

# ── Load data ──────────────────────────────────────────────────────────────

listener_csvs = glob_one(
    os.path.join(ROOT, "Apple", "Apple Monthly Listeners", "*.csv"),
    "Apple Monthly Listeners CSVs"
)
apple_df = pd.concat(
    [pd.read_csv(f, quotechar='"') for f in sorted(listener_csvs)],
    ignore_index=True
)
apple_df['Date'] = pd.to_datetime(apple_df['Date'].astype(str), format='%Y%m%d')
for col in ['Total Time Listened', 'Plays', 'Unique Listeners', 'Unique Engaged Listeners']:
    apple_df[col] = pd.to_numeric(apple_df[col])
apple_df = apple_df.sort_values('Date').reset_index(drop=True)

sub_csvs = glob_one(
    os.path.join(ROOT, "Apple", "Apple Monthly Subscribers", "*.csv"),
    "Apple Monthly Subscribers CSVs"
)
apple_subs_df = pd.concat(
    [pd.read_csv(f) for f in sorted(sub_csvs)],
    ignore_index=True
)
apple_subs_df['Date'] = pd.to_datetime(apple_subs_df['Date'].astype(str), format='%Y%m%d')
apple_subs_df = apple_subs_df.sort_values('Date').reset_index(drop=True)

spotify_files = glob_one(
    os.path.join(ROOT, "Spotify", "*Spotify_Listeners*.csv"),
    "Spotify Listeners CSV"
)
spotify_df = pd.read_csv(spotify_files[0])
spotify_df['Date'] = pd.to_datetime(spotify_df['Date'])
spotify_df = spotify_df.sort_values('Date').reset_index(drop=True)

viewer_files = glob_one(
    os.path.join(ROOT, "YouTube", "*AllViewers*.csv"),
    "YouTube AllViewers CSV"
)
yt_viewers = pd.read_csv(viewer_files[0])
yt_viewers['Date'] = pd.to_datetime(yt_viewers['Date'])
yt_viewers = yt_viewers.sort_values('Date').reset_index(drop=True)

sub_files = glob_one(
    os.path.join(ROOT, "YouTube", "*Subscribers*.csv"),
    "YouTube Subscribers CSV"
)
yt_subs = pd.read_csv(sub_files[0])
yt_subs['Date'] = pd.to_datetime(yt_subs['Date'])
yt_subs = yt_subs.sort_values('Date').reset_index(drop=True)

# ── Chart helpers ──────────────────────────────────────────────────────────

BG        = '#F2EDE3'
DARK_RED  = '#7B1E1E'
GRID_C    = '#C5BDB3'
TEXT_C    = '#3A3A3A'
APPLE_C   = '#5B2C8D'
SPOTIFY_C = '#1A4B8C'
YOUTUBE_C = '#C0392B'

def style_ax(ax, fig):
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG)
    for s in ['top', 'right', 'left']:
        ax.spines[s].set_visible(False)
    ax.spines['bottom'].set_color(GRID_C)
    ax.tick_params(colors=TEXT_C, labelsize=9)
    ax.yaxis.grid(True, color=GRID_C, linewidth=0.8)
    ax.set_axisbelow(True)

def weekly_xaxis(ax):
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=0))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%d-%b'))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right', fontsize=8)

def comma_fmt(ax):
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x):,}'))

def save_chart(fig, name):
    fig.savefig(os.path.join(IMG_DIR, name), dpi=150, bbox_inches='tight', facecolor=BG)
    plt.close(fig)

# ── Generate charts ────────────────────────────────────────────────────────

# Followers overview
fig, ax = plt.subplots(figsize=(13, 5.5))
style_ax(ax, fig)
ax.plot(apple_subs_df['Date'], apple_subs_df['Net Followers'], color=APPLE_C,   lw=1.8, label='Apple Podcasts')
ax.plot(spotify_df['Date'],    spotify_df['Followers'],        color=SPOTIFY_C, lw=1.8, label='Spotify')
ax.plot(yt_subs['Date'],       yt_subs['Subscribers'],         color=YOUTUBE_C, lw=1.8, label='YouTube')
weekly_xaxis(ax)
comma_fmt(ax)
ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.22), ncol=3,
          frameon=False, fontsize=9, labelcolor=TEXT_C)
plt.tight_layout()
save_chart(fig, 'chart_combined.png')

# Apple listeners
fig, ax = plt.subplots(figsize=(13, 5.5))
style_ax(ax, fig)
ax.fill_between(apple_df['Date'], apple_df['Unique Listeners'], color=DARK_RED, linewidth=0)
weekly_xaxis(ax)
comma_fmt(ax)
plt.tight_layout()
save_chart(fig, 'chart_apple.png')

# Spotify plays
fig, ax = plt.subplots(figsize=(13, 5.5))
style_ax(ax, fig)
ax.fill_between(spotify_df['Date'], spotify_df['Plays'], color=DARK_RED, linewidth=0)
weekly_xaxis(ax)
comma_fmt(ax)
plt.tight_layout()
save_chart(fig, 'chart_spotify.png')

# YouTube monthly audience
fig, ax = plt.subplots(figsize=(13, 5.5))
style_ax(ax, fig)
ax.fill_between(yt_viewers['Date'], yt_viewers['Monthly audience'], color=DARK_RED, linewidth=0)
weekly_xaxis(ax)
comma_fmt(ax)
plt.tight_layout()
save_chart(fig, 'chart_youtube.png')

# ── Build workbook ─────────────────────────────────────────────────────────

wb = Workbook()

HDR_FONT   = Font(name='Arial', bold=True, size=10, color='FFFFFF')
HDR_FILL   = PatternFill('solid', fgColor='2C3E50')
HDR_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(name='Arial', bold=True, size=14)
DATA_FONT  = Font(name='Arial', size=10)

def write_header(ws, row, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = HDR_ALIGN
    ws.row_dimensions[row].height = 30

def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def set_title(ws, text, merge_to):
    ws.row_dimensions[1].height = 24
    ws['A1'] = text
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A1:{merge_to}1')

def add_chart(ws, filename, anchor):
    img = XLImage(os.path.join(IMG_DIR, filename))
    img.anchor = anchor
    ws.add_image(img)

# Followers tab
ws_f = wb.active
ws_f.title = 'Followers'
set_title(ws_f, f'Machines Like Us – Followers & Subscribers  |  {PERIOD}', 'D')
write_header(ws_f, 3, ['Date', 'Apple Net Followers', 'Spotify Followers', 'YouTube Subscribers'])
set_col_widths(ws_f, {'A': 14, 'B': 24, 'C': 20, 'D': 22})

fol = (
    apple_subs_df[['Date', 'Net Followers']].rename(columns={'Net Followers': 'Apple'})
    .merge(spotify_df[['Date', 'Followers']].rename(columns={'Followers': 'Spotify'}), on='Date', how='outer')
    .merge(yt_subs[['Date', 'Subscribers']].rename(columns={'Subscribers': 'YouTube'}), on='Date', how='outer')
    .sort_values('Date').reset_index(drop=True)
)
for i, row in fol.iterrows():
    r = i + 4
    ws_f.cell(r, 1, row['Date'].date()).font = DATA_FONT
    ws_f.cell(r, 2, int(row['Apple'])   if pd.notna(row['Apple'])   else '').font = DATA_FONT
    ws_f.cell(r, 3, int(row['Spotify']) if pd.notna(row['Spotify']) else '').font = DATA_FONT
    ws_f.cell(r, 4, int(row['YouTube']) if pd.notna(row['YouTube']) else '').font = DATA_FONT
add_chart(ws_f, 'chart_combined.png', 'F2')

# Apple tab
ws_a = wb.create_sheet('Apple')
set_title(ws_a, f'Machines Like Us – Apple Podcasts Listeners  |  {PERIOD}', 'E')
write_header(ws_a, 3, ['Date', 'Total Time Listened', 'Plays', 'Unique Listeners', 'Unique Engaged Listeners'])
set_col_widths(ws_a, {'A': 14, 'B': 22, 'C': 10, 'D': 18, 'E': 24})
for i, row in apple_df.iterrows():
    r = i + 4
    ws_a.cell(r, 1, row['Date'].date()).font = DATA_FONT
    ws_a.cell(r, 2, int(row['Total Time Listened'])).font = DATA_FONT
    ws_a.cell(r, 3, int(row['Plays'])).font = DATA_FONT
    ws_a.cell(r, 4, int(row['Unique Listeners'])).font = DATA_FONT
    ws_a.cell(r, 5, int(row['Unique Engaged Listeners'])).font = DATA_FONT
add_chart(ws_a, 'chart_apple.png', 'G2')

# Spotify tab
ws_s = wb.create_sheet('Spotify')
set_title(ws_s, f'Machines Like Us – Spotify Listeners  |  {PERIOD}', 'D')
write_header(ws_s, 3, ['Date', 'Plays', 'Consumption Hours', 'Followers'])
set_col_widths(ws_s, {'A': 14, 'B': 10, 'C': 20, 'D': 12})
for i, row in spotify_df.iterrows():
    r = i + 4
    ws_s.cell(r, 1, row['Date'].date()).font = DATA_FONT
    ws_s.cell(r, 2, int(row['Plays'])).font = DATA_FONT
    ws_s.cell(r, 3, round(float(row['Consumption hours']), 2)).font = DATA_FONT
    ws_s.cell(r, 4, int(row['Followers'])).font = DATA_FONT
add_chart(ws_s, 'chart_spotify.png', 'F2')

# YouTube tab
ws_y = wb.create_sheet('YouTube')
set_title(ws_y, f'Machines Like Us – YouTube Viewers  |  {PERIOD}', 'C')
yt_merged = yt_viewers.merge(yt_subs, on='Date', how='outer').sort_values('Date').reset_index(drop=True)
write_header(ws_y, 3, ['Date', 'Monthly Audience', 'Subscribers'])
set_col_widths(ws_y, {'A': 14, 'B': 18, 'C': 14})
for i, row in yt_merged.iterrows():
    r = i + 4
    ws_y.cell(r, 1, row['Date'].date()).font = DATA_FONT
    ws_y.cell(r, 2, int(row['Monthly audience']) if pd.notna(row.get('Monthly audience')) else '').font = DATA_FONT
    ws_y.cell(r, 3, int(row['Subscribers'])       if pd.notna(row.get('Subscribers'))       else '').font = DATA_FONT
add_chart(ws_y, 'chart_youtube.png', 'E2')

# Save
period_slug = PERIOD.replace(' ', '_')
out_filename = f"MLU_{period_slug}_{REPORT_TYPE}.xlsx"
out_path = os.path.join(ROOT, out_filename)
wb.save(out_path)
print(f"SAVED:{out_path}")

# Cleanup chart images
shutil.rmtree(IMG_DIR, ignore_errors=True)
